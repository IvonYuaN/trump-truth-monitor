#!/usr/bin/env python3
"""
特朗普 Truth Social 每日监控 - 主运行脚本
"""

import os
import json
import requests
import re
import html
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== 配置 =====
DATA_URL = "https://ix.cnn.io/data/truth-social/truth_archive.json"
DATA_DIR = Path("data")
OUTPUT_DIR = Path("output")
ARCHIVE_FILE = DATA_DIR / "truth_archive.json"

# ===== 翻译词典 =====
KEY_TRANSLATIONS = {
    "NATO WASN'T THERE WHEN WE NEEDED THEM, AND THEY WON'T BE THERE IF WE NEED THEM AGAIN. REMEMBER GREENLAND, THAT BIG, POORLY RUN, PIECE OF ICE!!!":
        "北约在我们需要他们的时候不在，如果我们再次需要他们，他们也不会在。记住格陵兰岛，那块管理混乱的大冰原！！！",
    "Tonight, an entire civilization will die, never to return.":
        "今晚，整个文明将消亡，永远无法重建。",
    "TACO!": "TACO！（塔可——暗示达成临时协议）",
    "Iran must pay for regional aggression.": "伊朗必须为其地区侵略行为付出代价。",
    "I'M BACK!": "我回来了！",
    "We found him!": "我们找到他了！",
    "Tuesday, 8 PM Eastern Time!": "周二，美东时间晚上8点！",
}

PHRASE_DICT = {
    "America First": "美国优先",
    "MAGA Warrior": "MAGA战士",
    "Complete and Total Endorsement": "完全且彻底的背书",
    "HE WILL NEVER LET YOU DOWN": "他绝不会让你们失望",
    "SHE WILL NEVER LET YOU DOWN": "她绝不会让你们失望",
    "President DJT": "总统 DJT",
    "RINO": "RINO（名义上的共和党人）",
    "Second Amendment": "第二修正案",
    "MADE IN THE U.S.A.": "美国制造",
    "American Energy DOMINANCE": "美国能源主导地位",
    "LAW AND ORDER": "法律与秩序",
    "PEACE THROUGH STRENGTH": "以实力求和平",
    "Stop Migrant Crime": "制止移民犯罪",
    "Cut Taxes and Regulations": "削减税收和监管",
    "Grow our Economy": "发展我们的经济",
    "Strait of Hormuz": "霍尔木兹海峡",
    "is doing a fantastic job": "做得非常出色",
    "is doing a tremendous job": "做得非常出色",
    "Congressional District": "国会选区",
    "State Senate District": "州参议院选区",
    "for Re-Election": "竞选连任",
    "It is my Great Honor to endorse": "我非常荣幸地背书",
}

def clean_html(text):
    if not text:
        return ""
    text = html.unescape(text)
    text = re.sub(r'<[^>]+>', '', text)
    text = text.replace('\xa0', ' ').strip()
    return text

def translate_post(text):
    if not text:
        return "[仅媒体/无文字内容]"
    result = text
    for en, cn in KEY_TRANSLATIONS.items():
        if en in result:
            result = result.replace(en, cn)
    sorted_phrases = sorted(PHRASE_DICT.items(), key=lambda x: len(x[0]), reverse=True)
    for en, cn in sorted_phrases:
        result = result.replace(en, cn)
    return result.strip()

def fetch_data():
    """获取数据"""
    print(f"[{datetime.now()}] 正在获取数据...")
    try:
        response = requests.get(DATA_URL, timeout=60, verify=False)
        response.raise_for_status()
        data = response.json()
        print(f"获取到 {len(data)} 条帖子")
        return data
    except Exception as e:
        print(f"获取数据失败: {e}")
        return None

def load_archive():
    """加载存档"""
    if ARCHIVE_FILE.exists():
        with open(ARCHIVE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_archive(data):
    """保存存档"""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(ARCHIVE_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"存档已保存: {ARCHIVE_FILE}")

def create_excel(posts, filename):
    """生成 Excel"""
    # 样式
    HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    TITLE_FONT = Font(bold=True, size=14, name="Arial", color="1F4E79")
    DATA_FONT = Font(size=10, name="Arial", color="333333")
    CN_FONT = Font(size=10, name="Microsoft YaHei", color="333333")
    DATE_FONT = Font(size=10, name="Arial", color="1F4E79", bold=True)
    NUM_FONT = Font(size=9, name="Arial", color="888888")
    LINK_FONT = Font(size=9, name="Arial", color="4472C4", underline="single")
    ZEBRA_FILL_1 = PatternFill(fill_type="solid", fgColor="FFFFFF")
    ZEBRA_FILL_2 = PatternFill(fill_type="solid", fgColor="F2F7FB")
    THIN_BORDER = Border(
        left=Side(style="thin", color="D9DEE7"),
        right=Side(style="thin", color="D9DEE7"),
        top=Side(style="thin", color="D9DEE7"),
        bottom=Side(style="thin", color="D9DEE7"),
    )
    WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "全部帖子"
    
    # 标题
    ws.merge_cells("A1:H1")
    ws["A1"] = f"特朗普 Truth Social 帖子整理 · {datetime.now().strftime('%Y年%m月%d日')} · 共 {len(posts)} 条"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # 表头
    headers = ["序号", "发布时间", "英文原文", "中文翻译", "链接", "回复", "转发", "点赞"]
    widths = [6, 16, 50, 50, 8, 7, 7, 7]
    
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    
    # 数据
    for i, post in enumerate(posts):
        r = 4 + i
        content = clean_html(post.get("content", ""))
        cn = translate_post(content)
        
        ws.cell(r, 1, i+1).font = NUM_FONT
        ws.cell(r, 1).alignment = CENTER
        
        ws.cell(r, 2, post["created_at"][:16].replace("T", " ")).font = DATE_FONT
        ws.cell(r, 2).alignment = CENTER
        
        ws.cell(r, 3, content if content else "[仅媒体]").font = DATA_FONT
        ws.cell(r, 3).alignment = WRAP
        
        ws.cell(r, 4, cn).font = CN_FONT
        ws.cell(r, 4).alignment = WRAP
        
        ws.cell(r, 5, "查看").font = LINK_FONT
        ws.cell(r, 5).alignment = CENTER
        if post.get("url"):
            ws.cell(r, 5).hyperlink = post["url"]
        
        ws.cell(r, 6, post.get("replies_count", 0)).font = NUM_FONT
        ws.cell(r, 6).alignment = CENTER
        ws.cell(r, 7, post.get("reblogs_count", 0)).font = NUM_FONT
        ws.cell(r, 7).alignment = CENTER
        ws.cell(r, 8, post.get("favourites_count", 0)).font = NUM_FONT
        ws.cell(r, 8).alignment = CENTER
        
        fill = ZEBRA_FILL_1 if i % 2 == 0 else ZEBRA_FILL_2
        for c in range(1, 9):
            ws.cell(r, c).fill = fill
            ws.cell(r, c).border = THIN_BORDER
        
        ws.row_dimensions[r].height = 40
    
    ws.freeze_panes = "A4"
    
    # 保存
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    print(f"Excel 已保存: {filepath}")
    return filepath

def main():
    # 获取数据
    all_posts = fetch_data()
    if not all_posts:
        print("无法获取数据，退出")
        return
    
    # 加载存档
    existing = load_archive()
    existing_ids = {p['id'] for p in existing}
    
    # 检查新帖子
    new_posts = [p for p in all_posts if p['id'] not in existing_ids]
    print(f"发现 {len(new_posts)} 条新帖子")
    
    # 合并存档
    all_posts_sorted = sorted(all_posts, key=lambda x: x['created_at'], reverse=True)
    save_archive(all_posts_sorted)
    
    # 筛选最近30天
    cutoff = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    recent_posts = [p for p in all_posts_sorted if p['created_at'] >= cutoff]
    
    # 今日帖子
    today = datetime.now().strftime("%Y-%m-%d")
    today_posts = [p for p in all_posts_sorted if p['created_at'].startswith(today)]
    
    # 生成 Excel
    filename = f"特朗普Truth_Social帖子_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = create_excel(recent_posts, filename)
    
    # 输出统计
    print(f"\n{'='*50}")
    print(f"📊 统计信息:")
    print(f"   总帖子数: {len(all_posts_sorted)}")
    print(f"   最近30天: {len(recent_posts)}")
    print(f"   今日新帖: {len(today_posts)}")
    print(f"{'='*50}")
    
    # 设置环境变量供后续使用
    with open(OUTPUT_DIR / "stats.txt", "w") as f:
        f.write(f"TOTAL_POSTS={len(all_posts_sorted)}\n")
        f.write(f"RECENT_POSTS={len(recent_posts)}\n")
        f.write(f"TODAY_POSTS={len(today_posts)}\n")
        f.write(f"NEW_POSTS={len(new_posts)}\n")
    
    return {
        "filepath": str(filepath),
        "total_posts": len(all_posts_sorted),
        "recent_posts": len(recent_posts),
        "today_posts": len(today_posts),
        "new_posts": len(new_posts),
    }

if __name__ == "__main__":
    main()
